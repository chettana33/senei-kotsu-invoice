#!/usr/bin/env python3
"""king-bus-taicho — 台帳 workflow บน Google Sheets (ไฟล์หลัก).

Subcommands:
  status                 - โชว์台帳เดือน 8-10 จาก Google Sheets
  diff  --form <xlsx>    - เทียบใบขอรถล่าสุดกับ台帳 (แสดงสิ่งที่ต้องลง/แก้/ลบ)
  apply --form <xlsx>    - เขียนผล diff ลง Google Sheets จริง
  pdf   --date <MMDD>    - export PDF แนวนอน A4 ช่วง A14:AG59 (ชื่อตามแพตเทิร์น MMDD)

ตัวอย่าง:
  python tools/taicho_gsheets.py status
  python tools/taicho_gsheets.py diff --form "H:\\...\\8月\\0817_....xlsx"
  python tools/taicho_gsheets.py apply --form "H:\\...\\8月\\0817_....xlsx"
  python tools/taicho_gsheets.py pdf --date 0817
"""
import argparse
import datetime
import io
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request

import openpyxl

SHEET_ID = "1H2WE2D8ZXrAI4jdOUm2N6DYGCWVD1SrYy9BqAfRFdC0"  # master จริง (พี่เจยืนยัน 29 ส.ค. 69) — เดิม 1LSQqblDx... ผิดไฟล์
SHEET_NAME = "千栄1568 9月-12月(2026)"  # ตาราง 4 เดือนเลื่อน (9-12, 10-1, ...) — rename จาก 7月-10月 31 ส.ค. 69; ใช้ find_main_tab() หา tab จริงอัตโนมัติ
GID = "1801459155"
CRED = r"C:\Users\chett\.google_workspace_mcp\credentials\kimonoland.jp1@gmail.com.sheets.json"
FORM_BASE = r"H:\My Drive\チェー（個人）\King BUS\台帳入力（配車時間入力）\2026"

# tab ของใบขอรถ -> เดือน (SOP: ตรวจ 26.8-26.12 และ 27.01-27.03)
TAB_MONTHS = {
    "26.8": 8, "26.9": 9, "26.10": 10, "26.11": 11, "26.12": 12,
    "27.01": 1, "27.02": 2, "27.03": 3,
}


# ---------------- Google API ----------------

def get_token():
    with open(CRED) as f:
        c = json.load(f)
    body = urllib.parse.urlencode({
        "client_id": c["client_id"],
        "client_secret": c["client_secret"],
        "refresh_token": c["refresh_token"],
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=body)
    return json.load(urllib.request.urlopen(req, timeout=20))["access_token"]


def find_main_tab():
    """หาชื่อ tab 台帳 หลักแบบ dynamic: ชื่อ '千栄1568 <M0>月-<M3>月(<ปี>)' — rotate รายเดือนสร้าง tab ใหม่
    (tab เก่าเก็บไว้เป็นประวัติ) → เลือก tab ที่ M0 (เดือนแรก) มากสุด = tab ล่าสุด.
    TEST_TAICHO_TAB env = บังคับใช้ tab ที่ระบุ (ทดสอบ). คืน (title, gid) หรือ (None, None)."""
    tok = get_token()
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}?fields=sheets(properties(title,sheetId))"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {tok}"})
    d = json.load(urllib.request.urlopen(req, timeout=30))
    test = os.environ.get("TEST_TAICHO_TAB")
    best = None
    for s in d["sheets"]:
        t = s["properties"]["title"]
        if test:
            if t == test:
                return t, s["properties"]["sheetId"]
            continue
        m = re.match(r"^千栄1568 (\d+)月-\d+月", t)
        if m:
            m0 = int(m.group(1))
            if best is None or m0 > best[0]:
                best = (m0, t, s["properties"]["sheetId"])
    if best:
        return best[1], best[2]
    return None, None


def sheets_get(values_range):
    tok = get_token()
    name, _ = find_main_tab()
    if not name:
        raise RuntimeError("ไม่พบ tab 台帳 หลัก (千栄1568 <M>月-<M>月)")
    rng = urllib.parse.quote(f"'{name}'!{values_range}")
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{rng}?majorDimension=ROWS"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {tok}"})
    return json.load(urllib.request.urlopen(req, timeout=30))


def sheets_update(values_range, values):
    tok = get_token()
    name, _ = find_main_tab()
    if not name:
        raise RuntimeError("ไม่พบ tab 台帳 หลัก (千栄1568 <M>月-<M>月)")
    rng = urllib.parse.quote(f"'{name}'!{values_range}")
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{rng}?valueInputOption=USER_ENTERED"
    body = json.dumps({"values": values}).encode()
    req = urllib.request.Request(url, data=body, method="PUT",
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req, timeout=30))


def drive_upload(file_path):
    """อัปโหลดไฟล์ขึ้น Google Drive (multipart) — คืน file id หรือ None (29 ส.ค. 69)."""
    tok = get_token()
    name = os.path.basename(file_path)
    with open(file_path, "rb") as f:
        media = f.read()
    boundary = "----taicho" + str(int(time.time() * 1000))
    meta = json.dumps({"name": name}).encode("utf-8")
    body = (b"--" + boundary.encode() + b"\r\n"
            b"Content-Type: application/json; charset=UTF-8\r\n\r\n" + meta + b"\r\n"
            b"--" + boundary.encode() + b"\r\n"
            b"Content-Type: application/pdf\r\n\r\n" + media + b"\r\n"
            b"--" + boundary.encode() + b"--\r\n")
    url = "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart"
    req = urllib.request.Request(url, data=body, method="POST",
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": f"multipart/related; boundary={boundary}"})
    try:
        resp = json.load(urllib.request.urlopen(req, timeout=60))
        return resp.get("id")
    except Exception as e:
        print(f"drive_upload FAILED: {e}")
        return None


# ---------------- 台帳 layout ----------------

def month_sections(rows, base_row):
    """rows = ค่า A14 ลงมา; คืน dict {month: entry_row} (แถวจริงใน sheet)."""
    sections = {}
    for i, row in enumerate(rows):
        a = row[0] if row else None
        if a and isinstance(a, str):
            m = re.search(r"(\d+)月", a)
            if m:
                # header แถว = base_row + i; entry แถว = header + 3
                sections[int(m.group(1))] = base_row + i + 3
    return sections


def col_for_day(day):
    """คอลัมน์ Excel ของวันที่: B=1, C=2, ..., AF=31"""
    n = 1 + day  # day 1 -> col 2 (B)
    if n <= 26:
        return chr(ord("A") + n - 1)
    return "A" + chr(ord("A") + n - 27)


def read_taicho():
    # อ่านจาก A1 (ตาราง 4 เดือนเลื่อน: เดือนแรกอาจเริ่มแถว 3 หรือ 9 — ห้าม hardcode A14)
    res = sheets_get("A1:AG80")
    rows = res.get("values", [])
    sections = month_sections(rows, 1)
    taicho = {}
    for month, entry_row in sections.items():
        idx = entry_row - 1  # 1-based row -> 0-based index (base_row=1)
        entry = rows[idx] if idx < len(rows) else []
        cells = {}
        for day in range(1, 32):
            c = col_for_day(day)
            ci = ord(c[0]) - ord("A") if len(c) == 1 else (ord(c[0]) - ord("A")) * 26 + ord(c[1]) - ord("A") + 26
            v = entry[ci] if ci < len(entry) else None
            if v not in (None, ""):
                cells[day] = v
        # จำนวนวันจริง = คอลัมน์วันใน header row (เช่น ส.ค. ใน sheet มีแค่ 1-30)
        ndays = DAYS_IN_MONTH.get(month, 30)
        hidx = idx - 3
        if hidx >= 0:
            hdr = rows[hidx] if hidx < len(rows) else []
            nums = []
            for v in hdr[1:]:
                try:
                    n = int(v)
                    if 1 <= n <= 31:
                        nums.append(n)
                except (TypeError, ValueError):
                    pass
            if nums:
                ndays = max(nums)
        taicho[month] = {"row": entry_row, "cells": cells, "ndays": ndays}
    return taicho


# ---------------- ใบขอรถ ----------------

def serial_to_date(n):
    if isinstance(n, datetime.datetime):
        return n.date()
    if isinstance(n, (int, float)):
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)).date()
    return None


def fmt_time(c):
    """คอลัมน์ C -> เวลาที่ใช้ลง台帳 (str 'HH:MM')."""
    if isinstance(c, datetime.time):
        return c.strftime("%H:%M")
    if isinstance(c, str):
        # รูปแบบ '7:45\n→08:00' -> ใช้ค่าหลังลูกศร (เวลาใหม่)
        if "→" in c:
            c = c.split("→")[-1].strip()
        c = re.sub(r"\s+", "", c)
        if re.match(r"^\d{1,2}[:：]\d{2}$", c):
            hh, mm = re.split(r"[:：]", c)
            return f"{int(hh):02d}:{mm}"
    return None


def hours_from(keiyu):
    """'AEON(2H)' -> '2H' / '成田山(1H)' -> '1H' / None"""
    if keiyu:
        m = re.search(r"\((\d+)H\)", str(keiyu))
        if m:
            return f"{m.group(1)}H"
    return None


# D2 ทิศทาง (mirror cloud main.py — พี่เจอนุมัติ 5 ก.ย. 69): ซ้าย=出発 ขวา=行先, 空=ฟ้า ホ=ส้ม
ROUTE_CODE = {"HTL": "ホ", "AP1": "空", "AP2": "空"}
C_BLUE = {"red": 0.12, "green": 0.38, "blue": 0.85}
C_ORANGE = {"red": 0.87, "green": 0.42, "blue": 0.06}


def col_index(col):
    """'A'->0 ... 'Z'->25, 'AA'->26 ..."""
    if len(col) == 1:
        return ord(col[0]) - 65
    return (ord(col[0]) - 65) * 26 + (ord(col[1]) - 65) + 26


def d2_text(entry):
    """entry {t, h?, D, G} -> 'ホ 10:00 2H 空' (ไม่มีรหัส = คงแบบเดิม '10:00 2H')"""
    o = ROUTE_CODE.get(entry.get("D", "")) or ""
    g2 = ROUTE_CODE.get(entry.get("G", "")) or ""
    t = entry["t"]
    h = entry.get("h") or ""
    mid = f"{t}" + (f" {h}" if h else "")
    return f"{o} {mid} {g2}" if (o and g2) else mid


def color_runs(text):
    """runs สี ทุก 空=ฟ้า / ホ=ส้ม (textFormatRuns = startIndex + format)"""
    runs = []
    off = 0
    for ch in text:
        n = len(ch.encode("utf-16-le")) // 2
        col = C_BLUE if ch == "空" else C_ORANGE if ch == "ホ" else None
        if col:
            runs.append({"startIndex": off, "format": {"foregroundColor": col}})
        off += n
    return runs


def cell_canon(text):
    """ชุด (เวลา, ชั่วโมง) ของข้อความหลายบรรทัด สำหรับ diff — ไม่สน 空/ホ/🔷/flag"""
    pairs = []
    for ln in (text or "").split("\n"):
        m = re.search(r"(\d{2}:\d{2})", ln)
        if not m:
            continue
        h = None
        hm = re.search(r"(\d+)H", ln)
        if hm:
            h = hm.group(1)
        pairs.append((m.group(1), h or ""))
    return "|".join(f"{t} {h}".strip() for t, h in sorted(pairs))


def write_cell_rich(entry_row, day, text, runs):
    """เขียนเซลล์วันพร้อม rich runs (สี D2) — ใช้ batchUpdate (runs ใส่กับ values PUT ไม่ได้)"""
    tok = get_token()
    name, gid = find_main_tab()
    col = col_for_day(day)
    ci = col_index(col)
    body = json.dumps({"requests": [{
        "updateCells": {
            "range": {"sheetId": gid, "startRowIndex": entry_row - 1, "startColumnIndex": ci,
                      "endRowIndex": entry_row, "endColumnIndex": ci + 1},
            "rows": [{"values": [{"userEnteredValue": {"stringValue": text},
                                  "textFormatRuns": runs}]}],
            "fields": "userEnteredValue,textFormatRuns",
        }}]}).encode()
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}:batchUpdate"
    req = urllib.request.Request(url, data=body, method="POST",
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": "application/json"})
    urllib.request.urlopen(req, timeout=30).read()


def read_request_form(path):
    """อ่านใบขอรถ: คืน dict {date: [entry]} entry = {t, h, D, G} (D=配車場所 col D, G=行先 col G)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for tab, month in TAB_MONTHS.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(10, ws.max_row + 1):
            n = ws.cell(row=r, column=14).value  # N = 配車日
            d = serial_to_date(n)
            if d is None or d.month != month or d.year not in (2026, 2027):
                continue
            j = ws.cell(row=r, column=10).value  # J = 連絡先/สถานะ
            if j and "キャンセル" in str(j):
                continue  # งานยกเลิก ไม่ลง台帳
            c = fmt_time(ws.cell(row=r, column=3).value)
            if c is None:
                continue
            e = ws.cell(row=r, column=5).value  # 経由地①
            f = ws.cell(row=r, column=6).value  # 経由地②
            h = hours_from(e) or hours_from(f)
            D = str(ws.cell(row=r, column=4).value or "").strip()  # 配車場所
            G = str(ws.cell(row=r, column=7).value or "").strip()  # 行先
            out.setdefault(d, []).append({"t": c, "h": h, "D": D, "G": G})
    for d in out:
        # เก็บเวลาซ้ำ (รถ 2 คันเวลาเดียวกัน) — set() ตัดซ้ำ ผิด (29 ส.ค. 69)
        out[d] = sorted(out[d], key=lambda e: (e["t"][:2], e["t"][3:]))
    return out


# ---------------- diff / apply ----------------

def strip_flags(s):
    """ตัดเครื่องหมายติดตาม 🟡/🟢 + newline ท้ายออก (ไม่นับเป็นค่าจริง)."""
    return re.sub(r"[🟡🟢]", "", s or "").rstrip("\r\n")


def build_plan(form_rows):
    """คืน dict {month: {day: (target_text, reason, cur)}} — diff แบบ D2-aware (เวลา+ชั่วโมง)"""
    taicho = read_taicho()
    plan = {}
    for d, entries in sorted(form_rows.items()):
        month = d.month
        if month not in taicho:
            continue  # เดือนนี้ยังไม่มี section ใน sheet — ข้าม
        target_text = "\n".join(d2_text(e) for e in entries)
        current = taicho.get(month, {}).get("cells", {}).get(d.day)
        cur_str = str(current) if current is not None else None
        if cell_canon(cur_str) != cell_canon(target_text):
            reason = "ลบ (งานยกเลิก/ไม่มีแล้ว)" if not entries else (
                "ลงใหม่" if cur_str is None else "แก้ไข")
            plan.setdefault(month, {})[d.day] = (target_text, reason, cur_str)
    return plan


def apply_plan(plan, dry_run=True):
    for month, days in sorted(plan.items()):
        entry_row = read_taicho()[month]["row"]
        for day in sorted(days):
            target_text, reason, cur = days[day]
            col = col_for_day(day)
            # 🟢 = อัปเดต/เพิ่มวันนี้ (ต่อท้ายเหมือนของเดิม) — เขียน rich text กันสี D2 หลุด
            text = target_text
            if text:
                text = text + " 🟢"
            if dry_run:
                print(f"  [{reason}] {month}月{day:02d} {col}{entry_row}: "
                      f"{cur!r} -> {text!r}")
            else:
                write_cell_rich(entry_row, day, text, color_runs(text))
                print(f"  [เขียน D2] {month}月{day:02d} {col}{entry_row}: {cur!r} -> {text!r}")
                time.sleep(0.4)


# ---------------- PDF (Google Sheets export — แทน HTML+Edge 1 ก.ย. 69) ----------------

EDGE = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"  # ไม่ใช้แล้ว เก็บไว้เผื่อ
DAYS_IN_MONTH = {8: 30, 9: 30, 10: 31}  # ส.ค. ใน sheet มี 30 คอลัมน์ (1-30)


def pdf_filename(mm, mmdd):
    """ชื่อไฟล์ PDF ตามเดือนไส้ใน (tab 4 เดือนเลื่อน) ไม่ใช่เดือนสร้าง:
    mm=9 → '9月-12月(0901)', mm=10 → '10月-01月(1001)', mm=1 → '1月-04月(0101)'.
    ปี = ปีของเดือนแรก (ส.ค.-ธ.ค. = 2026, ม.ค.-ก.ค. = 2027). (พี่เจแก้ชื่อ 1 ก.ย. 69)"""
    m3 = mm + 3
    if m3 > 12:
        m3 -= 12
    year = 2026 if mm >= 8 else 2027
    return f"{year}台帳 - 千栄1568 - {mm}月-{m3:02d}月({mmdd}).pdf"


def export_pdf_sheets(out_path, month_dir):
    """Export PDF จาก Google Sheets โดยตรง — ใช้ print layout ของ tab จริง
    (A4 landscape + หัวเดือนสี 30/31 วัน + เส้น/เนื้อหาครบ — เหมือนที่พี่เจปริ้น -Jay).

    เดิม: HTML+Edge headless ได้ PDF แนวตั้ง (612x792) + เนื้อหาหาย เปิดไม่ได้ (1 ก.ย. 69).
    ใช้เป็น fallback เมื่อ Excel COM path ล้ม (lesson #55)."""
    os.makedirs(month_dir, exist_ok=True)
    tok = get_token()
    _, gid = find_main_tab()
    params = urllib.parse.urlencode({
        "format": "pdf", "gid": gid, "size": "A4", "portrait": "false",
        "fitw": "true", "gridlines": "false", "printtitle": "false",
        "top_margin": "0.25", "bottom_margin": "0.25",
        "left_margin": "0.25", "right_margin": "0.25",
    })
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?{params}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {tok}"})
    data = urllib.request.urlopen(req, timeout=60).read()
    full = os.path.join(month_dir, out_path)
    with open(full, "wb") as f:
        f.write(data)
    print(f"PDF saved (sheets fallback): {full} ({len(data)} bytes)")
    return full


# emoji -> font ที่ Sheets export ใช้ embed emoji glyph (lesson #58: ไม่ต้องแทนด้วย ●◆
# — ตั้ง fontFamily "Segoe UI Emoji" ใน run → PDF ได้ emoji สีจริง 842x595 เหมือน -Jay)
EMOJI_FONT = "Segoe UI Emoji"
EMOJI_SYMBOL = ("🟡", "🟢", "🔷")


def _read_emoji_cells():
    """อ่านค่าทั้ง tab หลัก (spreadsheets.get — เฉพาะ userEnteredValue ไม่แพง)
    คืน dict {(row, col): ค่าเดิม} เฉพาะเซลล์ที่มี emoji. (lesson #55: batchGet HTML 400
    → ใช้ get+fields แทน)"""
    tok = get_token()
    name, gid = find_main_tab()
    rng = urllib.parse.quote(f"'{name}'!A1:AG929")
    url = (f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}"
           f"?ranges={rng}&fields=sheets(data(rowData(values(userEnteredValue))))")
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {tok}"})
    d = json.load(urllib.request.urlopen(req, timeout=30))
    found = {}
    rows = d["sheets"][0].get("data", [{}])[0].get("rowData", [])
    for i, rd in enumerate(rows, start=1):
        for j, v in enumerate(rd.get("values", []), start=1):
            sv = v.get("userEnteredValue", {}).get("stringValue")
            if sv and any(e in sv for e in EMOJI_SYMBOL):
                found[(i, j)] = sv
    return found


def _batch_update(requests):
    tok = get_token()
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}:batchUpdate"
    body = json.dumps({"requests": requests}).encode()
    req = urllib.request.Request(url, data=body, method="POST",
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req, timeout=60))


def _utf16_offsets(s):
    """คืน list ตำแหน่ง UTF-16 start index ของแต่ละตัวอักษร (emoji = surrogate pair = 2 units;
    textFormatRuns startIndex นับแบบ UTF-16 ไม่ใช่ Python len)."""
    offsets, i = [], 0
    for ch in s:
        offsets.append((i, ch))
        i += 2 if ord(ch) > 0xFFFF else 1
    return offsets


def export_pdf_rich(out_path, month_dir):
    """PDF ผ่าน Sheets API rich text runs (เส้นทางหลัก 1 ก.ย. 69 — lesson #58 ทางที่ดีที่สุด):
    1. หาเซลล์ emoji (🟡🟢🔷) ทั่ว tab
    2. batchUpdate textFormatRuns: run ของ emoji ใช้ fontFamily "Segoe UI Emoji" (ฟอนต์ระบบ
       มี glyph emoji) → Sheets export ฝัง emoji glyph ลง PDF — ได้ emoji สีจริง 842x595
       เหมือน -Jay (ไม่ต้องแทนด้วย ●◆; lesson #55/#57 ทางนั้น = สีท่วม/ไม่ใช่ emoji จริง)
    3. export PDF จาก Sheets
    4. revert กลับค่าเดิม (finally) — sheet ไม่เปลี่ยนถาวร

    คืน path PDF เต็ม หรือ raise (caller fallback)."""
    os.makedirs(month_dir, exist_ok=True)
    _, gid = find_main_tab()
    cells = _read_emoji_cells()
    if not cells:
        print("ไม่มีเซลล์ emoji — export ตรง")
        return export_pdf_sheets(out_path, month_dir)
    requests = []
    for (r, c), v in cells.items():
        # runs: เริ่ม default ที่ 0 → emoji run (font emoji) ที่ตำแหน่ง → default ต่อ
        runs = []
        prev_end = 0
        for pos, ch in _utf16_offsets(v):
            if ch in EMOJI_SYMBOL:
                if pos > prev_end:
                    runs.append({"startIndex": prev_end, "format": {}})
                runs.append({"startIndex": pos, "format": {"fontFamily": EMOJI_FONT}})
                prev_end = pos + (2 if ord(ch) > 0xFFFF else 1)
        if prev_end < len(v.encode("utf-16-le")) // 2:
            runs.append({"startIndex": prev_end, "format": {}})
        requests.append({
            "updateCells": {
                "range": {"sheetId": gid, "startRowIndex": r - 1, "startColumnIndex": c - 1,
                          "endRowIndex": r, "endColumnIndex": c},
                "rows": [{"values": [{
                    "userEnteredValue": {"stringValue": v},
                    "textFormatRuns": runs,
                }]}],
                "fields": "userEnteredValue,textFormatRuns",
            }
        })
    try:
        _batch_update(requests)
        print(f"emoji font runs เขียน {len(cells)} เซลล์ แล้ว export...")
        pdf = export_pdf_sheets(out_path, month_dir)
        return _embed_emoji_images(pdf)
    finally:
        rev = [{
            "updateCells": {
                "range": {"sheetId": gid, "startRowIndex": r - 1, "startColumnIndex": c - 1,
                          "endRowIndex": r, "endColumnIndex": c},
                "rows": [{"values": [{"userEnteredValue": {"stringValue": v}}]}],
                "fields": "userEnteredValue",
            }
        } for (r, c), v in cells.items()]
        try:
            _batch_update(rev)
            print(f"revert กลับ {len(cells)} เซลล์ เรียบร้อย")
        except Exception as e:
            print(f"REVERT FAILED: {e} — sheet ยังมี font run emoji อยู่!")


def _embed_emoji_images(pdf_path):
    """แทนที่ text emoji (🟡🟢🔷) ใน PDF ด้วยภาพ PNG ที่ render จาก Segoe UI Emoji (Windows)
    แล้วฝังลง PDF — viewer ทุกตัวเห็น emoji สี (Glyph ของ NotoColorEmoji ที่ Sheets ฝังชื่อไว้
    ไม่มีใน Edge/Adobe → □; ภาพ = ชัวร์ทุกที่, lesson #59)."""
    try:
        import fitz
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return pdf_path
    doc = fitz.open(pdf_path)
    page = doc[0]
    font = ImageFont.truetype(r"C:\Windows\Fonts\seguiemj.ttf", 128)
    targets = []
    for b in page.get_text("dict")["blocks"]:
        for l in b.get("lines", []):
            for s in l["spans"]:
                if any(ch in s["text"] for ch in EMOJI_SYMBOL):
                    targets.append((fitz.Rect(s["bbox"]), s["text"]))
    if not targets:
        doc.close()
        return pdf_path
    import io
    for rect, text in targets:
        page.add_redact_annot(rect + (-0.3, -0.3, 0.3, 0.3), fill=(1, 1, 1))
        img = Image.new("RGBA", (160, 160), (0, 0, 0, 0))
        d = ImageDraw.Draw(img)
        d.text((16, 8), text, font=font, embedded_color=True)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        page.insert_image(rect, stream=buf.getvalue())
    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
    doc.save(pdf_path, incremental=True, encryption=fitz.PDF_ENCRYPT_KEEP)
    doc.close()
    print(f"embed emoji images {len(targets)} จุด ลง PDF")
    return pdf_path


def export_pdf(out_path, month_dir):
    """เส้นทางหลัก = HTML + Edge headless print (lessons #60 — emoji 🟡🟢🔷 ฝัง Type3
    เหมือน -Jay, ช่องกว้างตามเนื้อหา); ล้ม → Sheets API rich text runs → Sheets export ตรง."""
    try:
        import taicho_pdf  # lazy: taicho_pdf import taicho_gsheets (กัน circular)
        return taicho_pdf.build_pdf(out_path, month_dir)
    except Exception as e:
        print(f"html+edge path FAILED: {e} — fallback Sheets API rich text")
        try:
            return export_pdf_rich(out_path, month_dir)
        except Exception as e2:
            print(f"rich text path FAILED: {e2} — fallback plain Sheets export")
            return export_pdf_sheets(out_path, month_dir)


# ---------------- CLI ----------------

def main():
    ap = argparse.ArgumentParser(description="king-bus-taicho Google Sheets workflow")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("status")
    p.set_defaults(fn=cmd_status)

    p = sub.add_parser("diff")
    p.add_argument("--form", required=True, help="path ถึงใบขอรถ xlsx")
    p.set_defaults(fn=cmd_diff)

    p = sub.add_parser("apply")
    p.add_argument("--form", required=True)
    p.add_argument("--yes", action="store_true", help="ยืนยันเขียนจริง (ไม่ถาม)")
    p.set_defaults(fn=cmd_apply)

    p = sub.add_parser("pdf")
    p.add_argument("--date", required=True, help="MMDD เช่น 0817")
    p.add_argument("--month", help="MM (default: เดือนของ MMDD)")
    p.set_defaults(fn=cmd_pdf)

    args = ap.parse_args()
    args.fn(args)


def cmd_status(args):
    taicho = read_taicho()
    for month in sorted(taicho):
        info = taicho[month]
        print(f"=== {month}月 (row {info['row']}) ===")
        for day in sorted(info["cells"]):
            print(f"  {month}月{day:02d}: {info['cells'][day]!r}")


def cmd_diff(args):
    form = read_request_form(args.form)
    plan = build_plan(form)
    if not plan:
        print("ไม่มีความต่าง — 台帳ตรงกับใบขอรถแล้ว")
        return
    for month, days in sorted(plan.items()):
        print(f"--- {month}月 ---")
        for day in sorted(days):
            target, reason, cur = days[day]
            print(f"  [{reason}] {month}月{day:02d}: {cur!r} -> {target!r}")
    print()
    print(f"ทั้งหมด {sum(len(d) for d in plan.values())} จุด")


def cmd_apply(args):
    form = read_request_form(args.form)
    plan = build_plan(form)
    if not plan:
        print("ไม่มีความต่าง — ไม่ต้องเขียน")
        return
    print("แผนที่จะเขียน:")
    apply_plan(plan, dry_run=True)
    if not args.yes:
        ans = input("เขียนลง Google Sheets จริงหรือไม่? [y/N] ").strip().lower()
        if ans not in ("y", "yes"):
            print("ยกเลิก")
            return
    apply_plan(plan, dry_run=False)


def cmd_pdf(args):
    d = args.date
    mm = int(args.month or d[0:2])  # 2 ตัวแรก = เดือน (MMDD); เดิม d[2:4] = วัน ผิด (31月)
    name = pdf_filename(mm, d)
    month_dir = os.path.join(FORM_BASE, f"{mm}月")
    os.makedirs(month_dir, exist_ok=True)
    export_pdf(name, month_dir)


if __name__ == "__main__":
    main()
