#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
invoice_edenstar.py — สร้างใบแจ้งหนี้ ザ エディスターホテル成田 (前半/後半) อัตโนมัติ

งาน ザ エディスターホテル成田 - 請求書 (ส่วนหนึ่งของ king-bus-taicho)
รอบงาน: วันที่ 16 ของเดือน = 前半 (1-15) / วันที่ 1 ของเดือน = 後半 (16-สิ้นเดือน)

สิ่งที่ทำ (Phase 1 — ตาม SOP king-bus-taicho):
  1) อ่านใบขอรถล่าสุด sheet `26.<MM>` → กรองคอลัมน์ 団体（お客様）名 ตามช่วงวัน (เฉพาะรายการยืนยันแล้ว)
  2) สร้างไฟล์ invoice จาก template 7月 (แท็บ/NO/กำหนดชำระ/เส้นตาราง: กรอบนอกหนา เส้นในบาง)
  3) ร่างเมล์ (draft) ถึง fmo@ixam.jp + แนบไฟล์ — สร้างแค่ draft ไม่ส่งเอง
  4) แจ้ง Discord #ai-activity-log

ใช้งาน:
    python tools/invoice_edenstar.py --month 8 --half front            # 8月前半
    python tools/invoice_edenstar.py --month 8 --half back             # 8月後半 (รอบ 1 ก.ย.)
    python tools/invoice_edenstar.py --month 8 --half front --dry-run  # แสดงแผน ไม่เขียนไฟล์/ไม่ส่งอะไร
    python tools/invoice_edenstar.py --month 9 --half front --no-email --no-discord  # เฉพาะไฟล์

ข้อห้าม (ตาม AGENTS.md / SOP):
  - ห้ามส่งเมล์จริง — สร้างแค่ draft ให้พี่เจตรวจแล้วกดส่งเอง
  - หลังพี่เจส่งเมล์แล้ว ต้องแจ้ง LINE ซ้ำ (ขั้นตอนแยก ตามที่พี่เจสั่ง)
  - ห้ามเก็บ credentials ลง repo — ใช้ไฟล์ creds นอก repo
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import json
import os
import re
import shutil
import sys
import urllib.parse
import urllib.request
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Border, Side

# ------------------------------------------------------------------ Config ---

DEFAULT_YEAR = 2026
DEFAULT_BASE = r"H:\My Drive\チェー（個人）\King BUS\台帳入力（配車時間入力）"
DEFAULT_TEMPLATE = os.path.join(DEFAULT_BASE, "2026", "7月", "ザ エディスターホテル成田 - 請求書 - 7月前半.xlsx")
CREDS_FILE = r"C:\Users\chett\.google_workspace_mcp\credentials\kimonoland.jp1@gmail.com.sheets.json"
TOKEN_URL = "https://oauth2.googleapis.com/token"
GMAIL_API = "https://gmail.googleapis.com/gmail/v1/users/me"
DISCORD_CONFIG = os.path.expanduser(r"~/.config/opencode/.discord_config.json")
MODEL_NAME = "พาเที่ยว (Buffy/Freebuff)"

SHEET_PREFIX = "26."   # sheet ในใบขอรถ เช่น 26.8 = ส.ค. 2026

# ห้ามส่ง Discord เด็ดขาด (pattern ข้อมูลลับ)
SENSITIVE_PATTERNS = [
    r"ghp_[A-Za-z0-9]{36}", r"gho_[A-Za-z0-9]{36}", r"sk-[A-Za-z0-9]{20,}",
    r"AIza[A-Za-z0-9_-]{30,}", r"ya29\.[A-Za-z0-9_-]+",
    r"password\s*[:=]", r"token\s*[:=]", r"secret\s*[:=]", r"api[_-]?key\s*[:=]", r"otp\s*[:=]",
]


def thai_half(half: str) -> str:
    return "前半" if half == "front" else "後半"


def tab_number(half: str) -> int:
    return 1 if half == "front" else 2


def invoice_no(year: int, month: int, half: str) -> str:
    """NO = YYYYMM15-<เลขแท็บ> เช่น 20260815-1 (前半) / 20260815-2 (後半)"""
    return f"{year}{month:02d}15-{tab_number(half)}"


def payment_deadline(year: int, month: int, half: str) -> dt.date:
    """กฎ: 1-15 → สิ้นเดือนนั้น / 16-สิ้นเดือน → วันที่ 15 เดือนถัดไป"""
    if half == "front":
        return dt.date(year, month, calendar.monthrange(year, month)[1])
    next_month = month + 1
    next_year = year + (1 if next_month == 13 else 0)
    if next_month == 13:
        next_month = 1
    return dt.date(next_year, next_month, 15)


def sheet_name(month: int) -> str:
    """sheet ในใบขอรถ: ลอง 26.08 ก่อน แล้วค่อย 26.8 (ชื่อไม่แน่นอน)"""
    return [f"{SHEET_PREFIX}{month:02d}", f"{SHEET_PREFIX}{month}"]


def excel_serial_to_date(serial) -> dt.date:
    """Excel serial → date (epoch 1899-12-30)"""
    return (dt.datetime(1899, 12, 30) + dt.timedelta(days=int(serial))).date()


# ------------------------------------------------------------ ใบขอรถ --------

def find_latest_form(month_dir: str) -> str:
    """หาถิดขอรถล่าสุด: ชื่อขึ้นต้น MMDD (เช่น 0814_千栄交通...) เลือก prefix สูงสุด,
    ถ้าซ้ำกันเลือก mtime ล่าสุด (ข้ามไฟล์ ~$, .gsheet, 台帳, invoice)"""
    candidates = []
    if not os.path.isdir(month_dir):
        raise FileNotFoundError(f"ไม่พบโฟลเดอร์ {month_dir}")
    for name in os.listdir(month_dir):
        if not name.lower().endswith(".xlsx"):
            continue
        if name.startswith("~$") or name.startswith("ザ エディスターホテル") or "台帳" in name:
            continue
        m = re.match(r"(\d{4})", name)
        if m:
            candidates.append((int(m.group(1)), os.path.getmtime(os.path.join(month_dir, name)), name))
    if not candidates:
        raise FileNotFoundError(f"ไม่พบไฟล์ใบขอรถ (千栄交通㈱_貸切バス手配依頼書) ใน {month_dir}")
    candidates.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return os.path.join(month_dir, candidates[0][2])


def read_month_sheet(form_path: str, year: int, month: int):
    """อ่าน sheet 26.<MM> ในใบขอรถ → คืนรายการ (date, 団体名) เฉพาะรายการยืนยันแล้ว
    กฎ (SOP): ใช้คอลัมน์ B (団体（お客様）名) + N (配車日) — ห้ามเดาวันที่จากชื่อ"""
    wb = openpyxl.load_workbook(form_path, data_only=True)
    target = None
    for cand in sheet_name(month):
        if cand in wb.sheetnames:
            target = cand
            break
    if not target:
        avail = [s for s in wb.sheetnames if s.startswith(SHEET_PREFIX)]
        raise RuntimeError(f"ไม่พบ sheet {sheet_name(month)} ใน {os.path.basename(form_path)} (มี: {avail})")
    ws = wb[target]

    # หาแถวหัวตาราง (มี 団体（お客様）名 + 配車日)
    header_row = None
    for r in range(1, 20):
        b = ws.cell(row=r, column=2).value
        n = ws.cell(row=r, column=14).value
        if b and "団体" in str(b) and n and "配車" in str(n):
            header_row = r
            break
    if header_row is None:
        raise RuntimeError(f"หาแถวหัวตารางไม่เจอใน sheet {target}")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        group = ws.cell(row=r, column=2).value          # B = 団体（お客様）名
        time = ws.cell(row=r, column=3).value           # C = 配車時間
        date_val = ws.cell(row=r, column=14).value      # N = 配車日
        if not group or not str(group).strip():
            continue
        # วันที่: datetime / serial / string
        d = None
        if isinstance(date_val, (dt.datetime, dt.date)):
            d = date_val.date() if isinstance(date_val, dt.datetime) else date_val
        elif isinstance(date_val, (int, float)):
            d = excel_serial_to_date(date_val)
        elif isinstance(date_val, str) and date_val.strip():
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
                try:
                    d = dt.datetime.strptime(date_val.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        if d is None:
            print(f"  ⚠️ ข้ามแถว {r}: อ่านวันที่ไม่ได้ (ค่า={date_val!r})")
            continue
        if time is None:  # ยังไม่ยืนยัน (ไม่มีเวลา) — ไม่เอาเข้าตาราง
            print(f"  ⚠️ ข้ามแถว {r}: {group!r} ยังไม่มีเวลา (ไม่ยืนยัน)")
            continue
        # ชื่อกลุ่ม: ตัด newline ออกให้เหลือบรรทัดเดียว (ตามไฟล์ที่พี่เจตรวจแล้ว)
        rows.append((d, str(group).strip().replace("\r\n", "").replace("\r", "").replace("\n", "")))
    return rows


def filter_items(rows, year: int, month: int, half: str):
    """กรองตามช่วง: 前半 = วันที่ 1-15 / 後半 = 16-สิ้นเดือน เรียงตามวันที่"""
    if half == "front":
        items = [(d, g) for d, g in rows if d.year == year and d.month == month and 1 <= d.day <= 15]
    else:
        last = calendar.monthrange(year, month)[1]
        items = [(d, g) for d, g in rows if d.year == year and d.month == month and 16 <= d.day <= last]
    items.sort(key=lambda t: t[0])
    return items


# ------------------------------------------------------------- Invoice ------

def apply_borders(ws):
    """เส้นตาราง: กรอบนอกหนา (medium) + เส้นแบ่งในบาง (thin) — แถว 16-39 คอลัมน์ B-L
    หมายเหตุ: openpyxl เขียนเส้นซ้าย/ขวาของช่องใน merge (K:L) ไม่ได้ → ตั้งขวาที่ K (anchor) แล้ว L สะท้อนตาม"""
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    cols = [openpyxl.utils.get_column_letter(n) for n in range(2, 13)]  # B..L
    for r in range(16, 40):
        for col in cols:
            left, right, top, bottom = thin, thin, thin, thin
            if r == 16:
                top = medium
            if r == 39:
                bottom = medium
            if col == "B":
                left = medium
            if col == "K":
                right = medium  # anchor ของ merge K:L → L จะสะท้อนตาม
            ws[f"{col}{r}"].border = Border(left=left, right=right, top=top, bottom=bottom)


def build_invoice(template: str, out_path: str, year: int, month: int, half: str, items):
    """คัดลอก template 7月 → ตั้งค่า (แท็บ/NO/กำหนดชำระ/รายการ/เส้น) → เซฟไฟล์ใหม่
    คงแถวว่างหลังรายการสุดท้ายไว้ (ตาม SOP)"""
    if not os.path.exists(template):
        raise FileNotFoundError(f"ไม่พบ template: {template}")
    shutil.copyfile(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws.title = f"ザ エディスターホテル成田 - 請求書 {year}-{month}({tab_number(half)})"
    ws["K3"] = invoice_no(year, month, half)
    # J4 = =TODAY() คงไว้ (ใช้วันส่งจริง)

    # ล้างรายการเดิม (แถว 17-34)
    for r in range(17, 35):
        for col in ("A", "B", "C", "G", "I"):
            ws[f"{col}{r}"] = None

    # เขียนรายการ (พื้นที่ปกติสูงสุด 18 แถว = 17-34; เกิน → เตือนขึ้นแท็บใหม่)
    if len(items) > 18:
        print(f"  ⚠️ รายการ {len(items)} แถว เกิน 18 — SOP: ให้ขึ้นแท็บใหม่ (ต้องทำต่อด้วยมือ)")
    for i, (d, group) in enumerate(items[:18]):
        r = 17 + i
        ws[f"A{r}"] = 1 if r == 17 else f"=1+A{r-1}"
        ws[f"B{r}"] = dt.datetime(d.year, d.month, d.day)
        ws[f"C{r}"] = group
        ws[f"G{r}"] = 1
        ws[f"I{r}"] = f"=SUM(G{r}*H{r})"

    # กำหนดชำระ (D43)
    ws["D43"] = dt.datetime.combine(payment_deadline(year, month, half), dt.time())

    # เส้นตาราง
    apply_borders(ws)

    wb.save(out_path)
    return out_path


# -------------------------------------------------------------- Discord -----

def send_discord(message: str) -> str:
    def is_sensitive(text):
        return any(re.search(p, text, re.IGNORECASE) for p in SENSITIVE_PATTERNS)
    if is_sensitive(message):
        return "Discord: NOT SENT — พบคำต้องสงสัย"
    webhook = None
    if os.path.exists(DISCORD_CONFIG):
        try:
            with open(DISCORD_CONFIG, encoding="utf-8") as fh:
                webhook = json.load(fh).get("webhook_url")
        except Exception:
            pass
    if not webhook:
        return "Discord: NOT SENT — ไม่พบ .discord_config.json"
    payload = json.dumps({"content": message}).encode("utf-8")
    req = urllib.request.Request(webhook, data=payload,
                                 headers={"Content-Type": "application/json", "User-Agent": "PhaTiao-Invoice"},
                                 method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return "Discord: SENT" if resp.status in (200, 204) else f"Discord: NOT SENT — HTTP {resp.status}"
    except Exception as exc:
        return f"Discord: NOT SENT — {type(exc).__name__}"


# ---------------------------------------------------------------- Gmail -----

def refresh_access_token(creds: dict) -> str:
    body = urllib.parse.urlencode({
        "client_id": creds["client_id"],
        "client_secret": creds["client_secret"],
        "refresh_token": creds["refresh_token"],
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request(TOKEN_URL, data=body,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        data = json.loads(resp.read().decode())
    if "access_token" not in data:
        raise RuntimeError(f"Token refresh failed: {data}")
    return data["access_token"]


def email_subject_body(year: int, month: int, half: str):
    """ข้อความเดิมที่เคยใช้ (ตามเมล์ 7月前半): subject ใช้ エデイ..., เนื้อหาใช้ エディ...
    body = ขึ้นบรรทัดใหม่ 4 บรรทัด (พี่เจสั่ง 31 ส.ค. 69 — รอบต่อไปใช้แบบนี้)"""
    half_jp = thai_half(half)  # 前半/後半
    subject = f"ザ エディスターホテル成田 - 請求書 - {year}年{month}月分{half_jp}"
    body = ("お世話になっております。 \n"
            f"ザ エディスターホテル成田 - 請求書 - {year}年{month}月{half_jp}です。\n"
            "添付ファイルを確認してください。 \n"
            "宜しくお願いします。")
    return subject, body


def create_gmail_draft(creds: dict, subject: str, body: str, attach_path: str) -> str:
    """สร้าง draft (ไม่ส่ง) — คืน draft id"""
    token = refresh_access_token(creds)

    msg = MIMEMultipart()
    msg["To"] = "fmo@ixam.jp"
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    fname = os.path.basename(attach_path)
    with open(attach_path, "rb") as fh:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", fname))
    msg.attach(part)

    raw = __import__("base64").urlsafe_b64encode(msg.as_bytes()).decode()
    payload = json.dumps({"message": {"raw": raw}}).encode()
    req = urllib.request.Request(f"{GMAIL_API}/drafts", data=payload, method="POST",
                                 headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        draft = json.loads(resp.read().decode())
    return draft.get("id", "")


# ------------------------------------------------------------------ Main -----

def main() -> int:
    parser = argparse.ArgumentParser(description="สร้าง invoice ザ エディスターホテル成田 (前半/後半) — Phase 1")
    parser.add_argument("--month", type=int, required=True, help="เดือน 1-12")
    parser.add_argument("--half", choices=["front", "back"], required=True, help="front=前半(1-15) / back=後半(16-สิ้นเดือน)")
    parser.add_argument("--year", type=int, default=DEFAULT_YEAR)
    parser.add_argument("--base", default=DEFAULT_BASE, help=f"โฟลเดอร์ฐาน (default: {DEFAULT_BASE})")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="ไฟล์ template 7月前半")
    parser.add_argument("--form", default=None, help="ระบุไฟล์ใบขอรถ (ไม่ระบุ = เลือกไฟล์ล่าสุดอัตโนมัติ)")
    parser.add_argument("--out-dir", default=None, help="โฟลเดอร์ output (default: <base>/<year>/<MM>月)")
    parser.add_argument("--no-email", action="store_true", help="ข้ามการร่างเมล์ (draft)")
    parser.add_argument("--no-discord", action="store_true", help="ข้ามการแจ้ง Discord")
    parser.add_argument("--skip-if-exists", action="store_true", help="ถ้าไฟล์ output มีอยู่แล้ว → ข้ามไม่ทับ + แจ้ง Discord (ใช้กับ scheduled task รอบอัตโนมัติ)")
    parser.add_argument("--dry-run", action="store_true", help="แสดงแผน ไม่เขียนไฟล์/ไม่สร้าง draft/ไม่ส่ง Discord")
    args = parser.parse_args()

    if not 1 <= args.month <= 12:
        parser.error("--month ต้องอยู่ระหว่าง 1-12")

    month_dir = os.path.join(args.base, str(args.year), f"{args.month}月")
    out_dir = args.out_dir or month_dir
    half_jp = thai_half(args.half)

    print("=" * 60)
    print(f"Invoice ザ エディスターホテル成田 — {args.year}年{args.month}月{half_jp}")
    print("=" * 60)

    # 1) อ่านใบขอรถ
    if args.form:
        form = args.form
    else:
        form = find_latest_form(month_dir)
    print(f"\n[1/4] อ่านใบขอรถ: {os.path.basename(form)}")
    rows = read_month_sheet(form, args.year, args.month)
    items = filter_items(rows, args.year, args.month, args.half)
    print(f"      รายการที่กรองได้ ({half_jp}): {len(items)} รายการ")
    for d, g in items:
        print(f"        - {d} | {g}")

    if not items:
        print("\n⚠️ ไม่มีรายการในช่วงนี้ — ตรวจข้อมูลก่อน (อาจยังไม่มีใบขอรถของช่วงนี้)")
        return 1

    # 2) สร้างไฟล์ invoice
    out_name = f"ザ エディスターホテル成田 - 請求書 - {args.month}月{half_jp}.xlsx"
    out_path = os.path.join(out_dir, out_name)
    # guard: ไฟล์มีอยู่แล้ว → ข้ามไม่ทับ (รอบอัตโนมัติ 16/1 — กันทับไฟล์ที่สร้าง/แก้มือแล้ว)
    if args.skip_if_exists and os.path.exists(out_path) and not args.dry_run:
        mtime = dt.datetime.fromtimestamp(os.path.getmtime(out_path)).strftime("%Y-%m-%d %H:%M")
        print(f"⏭️ ไฟล์มีอยู่แล้ว (สร้าง {mtime}) — ข้ามสร้าง ไม่ทับ: {out_path}")
        skip_msg = (f"⏭️ {MODEL_NAME}: ไฟล์ invoice {args.year}年{args.month}月{half_jp} มีอยู่แล้ว (สร้าง {mtime}) "
                    f"— ข้ามสร้าง (ไม่ทับ) พี่เจตรวจ/ส่งเอง: {os.path.basename(out_path)}")
        if args.no_discord:
            print("      (ข้าม Discord)")
        else:
            print(f"      {send_discord(skip_msg)}")
        return 0
    print(f"\n[2/4] สร้างไฟล์: {out_path}")
    print(f"      NO: {invoice_no(args.year, args.month, args.half)} | แท็บ: ({tab_number(args.half)}) | กำหนดชำระ: {payment_deadline(args.year, args.month, args.half)}")
    if args.dry_run:
        print("      (dry-run) ข้ามการเขียนไฟล์")
    else:
        os.makedirs(out_dir, exist_ok=True)
        build_invoice(args.template, out_path, args.year, args.month, args.half, items)
        print(f"      เขียนแล้ว: {out_path}")

    # 3) Discord
    print("\n[3/4] Discord #ai-activity-log ...")
    discord_msg = (f"📄 {MODEL_NAME}: สร้าง invoice {args.year}年{args.month}月{half_jp} "
                   f"({len(items)} รายการ) → draft เมล์รอพี่เจตรวจ/ส่ง")
    if args.dry_run or args.no_discord:
        print("      (ข้าม)")
        discord_status = "Discord: SKIPPED"
    else:
        discord_status = send_discord(discord_msg)
        print(f"      {discord_status}")

    # 4) ร่างเมล์
    print("\n[4/4] ร่างเมล์ (draft, ไม่ส่ง) ...")
    subject, body = email_subject_body(args.year, args.month, args.half)
    print(f"      To: fmo@ixam.jp")
    print(f"      Subject: {subject}")
    print(f"      Body: {body}")
    if args.dry_run or args.no_email:
        print("      (ข้าม) สร้างแค่ draft ไม่ส่งเอง")
        draft_status = "Draft: SKIPPED"
    else:
        with open(CREDS_FILE, encoding="utf-8") as fh:
            creds = json.load(fh)
        draft_id = create_gmail_draft(creds, subject, body, out_path)
        draft_status = f"Draft: {draft_id}"
        print(f"      ✅ สร้าง draft แล้ว: {draft_id}")
        print(f"      ลิงก์ตรวจ/ส่ง: https://mail.google.com/mail/u/0/#drafts?compose={draft_id}")

    print("\n" + "=" * 60)
    print(f"ผลสรุป: {discord_status} | {draft_status}")
    print("⚠️ ยังไม่ส่งเมล์ — ให้พี่เจตรวจไฟล์แล้วกดส่งเอง / ส่งแล้วต้องแจ้ง LINE ซ้ำ")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
