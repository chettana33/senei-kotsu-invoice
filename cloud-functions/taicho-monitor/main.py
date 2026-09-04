# -*- coding: utf-8 -*-
"""
台帳 auto update — Cloud Function (Firebase gen2, HTTP) — cloud version ของ taicho_auto.py
Flow: Drive (ใบขอรถ xlsx ใหม่) -> diff 台帳 -> apply sheets -> LINE (text + webapp link)
       -> PDF archive (phase 2b: reportlab ล้วน — taicho_pdf_cloud.py — ไม่ต้อง browser)
PDF รายวัน: สร้างหลัง apply สำเร็จ (เช่น local เดิม) แล้วอัปโหลด Drive โฟลเดอร์เดือนเดียวกัน;
LINE ยังส่งลิงก์ Web App (เหมือน local — PDF = ไฟล์เก็บหลักฐาน)
State (processed + last_flag_date) = ไฟล์ JSON ใน Drive root (taicho_cloud_state.json)
Env: SHEETS_CLIENT_ID / SHEETS_CLIENT_SECRET / SHEETS_REFRESH_TOKEN (kimonoland sheets token,
     scope มี drive ด้วย — ใช้เรียก Drive API) / LINE_CHANNEL_ID / LINE_CHANNEL_SECRET / LINE_USER_IDS
"""
import base64
import io
import json
import os
import re
import tempfile
import time
import urllib.parse
import urllib.request
import datetime as _dt
from datetime import date

from firebase_functions import https_fn
from firebase_functions.options import SupportedRegion
import openpyxl

import taicho_pdf_cloud as tpc  # PDF renderer reportlab (phase 2b)

SHEET_ID = "1H2WE2D8ZXrAI4jdOUm2N6DYGCWVD1SrYy9BqAfRFdC0"
SHEETS_API = "https://sheets.googleapis.com/v4/spreadsheets"
DRIVE_API = "https://www.googleapis.com/drive/v3"
TOKEN_URL = "https://oauth2.googleapis.com/token"
LINE_TOKEN_URL = "https://api.line.me/v2/oauth/accessToken"
FORM_RE = re.compile(r"(\d{4})_千栄交通㈱_貸切バス手配依頼書(?:\(\d+\))?\.xlsx$")
TAB_MONTHS = {"26.8": 8, "26.9": 9, "26.10": 10, "26.11": 11, "26.12": 12,
              "27.01": 1, "27.02": 2, "27.03": 3}
DRIVE_PATH = ["チェー（個人）", "King BUS", "台帳入力（配車時間入力）"]
MONTH_BLOCK = 16
DAYS_IN_MONTH = {8: 30, 9: 30, 10: 31}
HEADER_COLOR = {1: (1, 0.9, 0.4), 2: (0.75, 0.75, 0.75), 3: (1, 0.65, 0.3),
                4: (0.6, 0.8, 1), 5: (1, 0.75, 0.8), 6: (0.7, 0.9, 0.7),
                7: (1, 0.9, 0.4), 8: (1, 0.65, 0.3), 9: (0.6, 0.8, 1),
                10: (1, 0.75, 0.8), 11: (0.7, 0.9, 0.7), 12: (1, 0.65, 0.3)}
STATE_NAME = "taicho_cloud_state.json"
TAICHO_WEBAPP_URL = ("https://script.google.com/macros/s/"
                     "AKfycbym-XXro-jSHP1yO2wgmV4RfNVSPyOFcJdiutRjSfTbEfwkODJxcDySGQ_oZN1rYQbc2A/exec")
SHORT_TAICHO_URL = "https://taicho-link.vercel.app"


def env(key):
    v = os.environ.get(key)
    if not v:
        raise RuntimeError(f"missing env var: {key}")
    return v


# ---------------- tokens ----------------

def sheets_token():
    body = urllib.parse.urlencode({
        "client_id": env("SHEETS_CLIENT_ID"), "client_secret": env("SHEETS_CLIENT_SECRET"),
        "refresh_token": env("SHEETS_REFRESH_TOKEN"), "grant_type": "refresh_token"}).encode()
    req = urllib.request.Request(TOKEN_URL, data=body)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]


def _api(method, url, token=None, body=None, timeout=60):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token or sheets_token()}")
    if body is not None:
        req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
        return json.loads(raw) if raw else {}


def _get_media(url, token):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


# ---------------- LINE ----------------

def send_line(text):
    cfg_body = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": env("LINE_CHANNEL_ID"), "client_secret": env("LINE_CHANNEL_SECRET")}).encode()
    req = urllib.request.Request(LINE_TOKEN_URL, data=cfg_body)
    with urllib.request.urlopen(req, timeout=30) as r:
        lt = json.loads(r.read())["access_token"]
    for uid in env("LINE_USER_IDS").split(","):
        if not uid.strip():
            continue
        body = json.dumps({"to": uid.strip(),
                           "messages": [{"type": "text", "text": text}]}).encode()
        req = urllib.request.Request("https://api.line.me/v2/bot/message/push", data=body,
                                     headers={"Authorization": f"Bearer {lt}",
                                              "Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=30).read()
    print("  LINE sent:", text.splitlines()[0][:60], flush=True)


# ---------------- Drive: state + forms ----------------

def drive_find_folder(name, parent, token):
    q = f"name='{name}' and '{parent}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    r = _api("GET", f"{DRIVE_API}/files?q={urllib.parse.quote(q)}&fields=files(id,name)", token)
    return (r.get("files") or [{}])[0].get("id")


def drive_resolve(path_parts, token):
    parent = "root"
    for name in path_parts:
        found = drive_find_folder(name, parent, token)
        if not found:
            meta = json.dumps({"name": name, "mimeType": "application/vnd.google-apps.folder",
                               "parents": [parent]}).encode()
            req = urllib.request.Request(f"{DRIVE_API}/files?fields=id", data=meta, method="POST",
                                         headers={"Authorization": f"Bearer {token}",
                                                  "Content-Type": "application/json"})
            found = json.loads(urllib.request.urlopen(req, timeout=30).read()).get("id")
        parent = found
    return parent


def state_file_id(token):
    q = f"name='{STATE_NAME}' and trashed=false"
    r = _api("GET", f"{DRIVE_API}/files?q={urllib.parse.quote(q)}&fields=files(id)", token)
    files = r.get("files") or []
    if files:
        return files[0]["id"]
    meta = json.dumps({"name": STATE_NAME, "mimeType": "application/json"}).encode()
    req = urllib.request.Request(f"{DRIVE_API}/files?fields=id", data=meta, method="POST",
                                 headers={"Authorization": f"Bearer {token}",
                                          "Content-Type": "application/json"})
    return json.loads(urllib.request.urlopen(req, timeout=30).read()).get("id")


def load_state(token):
    fid = state_file_id(token)
    try:
        raw = _get_media(f"{DRIVE_API}/files/{fid}?alt=media", token)
        return json.loads(raw.decode("utf-8"))
    except Exception:
        return {"processed": []}


def save_state(state, token):
    fid = state_file_id(token)
    data = json.dumps(state, ensure_ascii=False).encode()
    # media upload ต้องใช้ upload domain (/upload/drive/v3) — drive host กับ uploadType=media
    # ตอบ 200 แต่ไม่เขียน (lessons #81) — state เลยไม่เคยบันทึกตั้งแต่ phase 2a
    req = urllib.request.Request(f"https://www.googleapis.com/upload/drive/v3/files/{fid}?uploadType=media",
                                 data=data, method="PATCH",
                                 headers={"Authorization": f"Bearer {token}",
                                          "Content-Type": "application/json"})
    urllib.request.urlopen(req, timeout=30).read()


# ---------------- PDF archive (phase 2b) ----------------

def pdf_filename(mm, mmdd):
    """ชื่อไฟล์ PDF ตามเดือนไส้ใน (ตรง tools/taicho_gsheets.py pdf_filename):
    mm=9 -> '9月-12月', mm=10 -> '10月-01月'; ปี = 2026 ถ้า mm>=8 else 2027."""
    m3 = mm + 3
    if m3 > 12:
        m3 -= 12
    year = 2026 if mm >= 8 else 2027
    return f"{year}台帳 - 千栄1568 - {mm}月-{m3:02d}月({mmdd}).pdf"


def _drive_find_file(name, parent, token):
    q = f"name='{name}' and '{parent}' in parents and trashed=false"
    r = _api("GET", f"{DRIVE_API}/files?q={urllib.parse.quote(q)}&fields=files(id,name)", token)
    return (r.get("files") or [{}])[0].get("id")


def _multipart(meta, media, boundary):
    head = (f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n"
            + meta + f"\r\n--{boundary}\r\nContent-Type: application/pdf\r\n\r\n").encode()
    tail = f"\r\n--{boundary}--\r\n".encode()
    return head + media + tail


def _drive_upload_pdf(name, parent, data, token):
    """สร้างไฟล์ PDF ในโฟลเดอร์ Drive (ไม่มี = create; มีชื่อซ้ำ = update media กันกองซ้ำ)"""
    fid = _drive_find_file(name, parent, token)
    if fid:
        # PATCH media บางที 400 transient หลัง create ใหม่ (race) — retry 1 ครั้ง
        for attempt in (1, 2):
            try:
                req = urllib.request.Request(
                    f"https://www.googleapis.com/upload/drive/v3/files/{fid}?uploadType=media",
                    data=data, method="PATCH",
                    headers={"Authorization": f"Bearer {token}",
                             "Content-Type": "application/pdf"})
                urllib.request.urlopen(req, timeout=60).read()
                return fid
            except urllib.error.HTTPError:
                if attempt == 1:
                    time.sleep(2)
                    continue
                raise
    boundary = "taicho_pdf_boundary_7f3a"
    meta = json.dumps({"name": name, "parents": [parent]})
    body = _multipart(meta, data, boundary)
    req = urllib.request.Request(
        f"https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&fields=id,name",
        data=body, method="POST",
        headers={"Authorization": f"Bearer {token}",
                 "Content-Type": f"multipart/form-data; boundary={boundary}"})
    return json.loads(urllib.request.urlopen(req, timeout=60).read()).get("id")


def archive_pdf(mmdd, token):
    """สร้าง PDF 台帳 (reportlab) จากสถานะ tab ปัจจุบัน → อัป Drive โฟลเดอร์เดือนของใบขอรถ
    (โฟลเดอร์เดียวกับที่ local เดิมเก็บ: .../台帳入力（配車時間入力）/<ปี>/<mm>月).
    ล้ม = raise (caller จับแล้วไม่บล็อก flow)"""
    mm = int(mmdd[0:2])
    folder = drive_resolve(DRIVE_PATH + [str(2026 if mm >= 8 else 2027), f"{mm}月"], token)
    fd, tmp = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    try:
        taicho = read_taicho()
        tpc.render_taicho_pdf(taicho, mmdd, tmp)
        with open(tmp, "rb") as f:
            data = f.read()
        name = pdf_filename(mm, mmdd)
        fid = _drive_upload_pdf(name, folder, data, token)
        print(f"PDF archived: {name} ({len(data)} bytes, drive id {fid})", flush=True)
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def list_forms(token):
    """คืน [(mmdd, file_id, name)] เรียง mmdd จากทุก folder เดือนที่เกี่ยวข้อง"""
    found = []
    for year, months in ((2026, [8, 9, 10, 11, 12]), (2027, [1, 2, 3])):
        for mm in months:
            try:
                folder = drive_resolve(DRIVE_PATH + [str(year), f"{mm}月"], token)
            except Exception:
                continue
            q = f"'{folder}' in parents and trashed=false"
            r = _api("GET", f"{DRIVE_API}/files?q={urllib.parse.quote(q)}&fields=files(id,name)", token)
            for f in r.get("files") or []:
                m = FORM_RE.match(f["name"])
                if m:
                    found.append((m.group(1), f["id"], f["name"]))
    return sorted(found, key=lambda t: t[0])


# ---------------- sheets 台帳 ----------------

def find_main_tab():
    d = _api("GET", f"{SHEETS_API}/{SHEET_ID}?fields=sheets(properties(title,sheetId))")
    best = None
    for s in d["sheets"]:
        m = re.match(r"^千栄1568 (\d+)月-\d+月", s["properties"]["title"])
        if m:
            m0 = int(m.group(1))
            if best is None or m0 > best[0]:
                best = (m0, s["properties"]["title"], s["properties"]["sheetId"])
    return (best[1], best[2]) if best else (None, None)


def sheets_get(values_range):
    name, _ = find_main_tab()
    if not name:
        raise RuntimeError("ไม่พบ tab 台帳 หลัก")
    rng = urllib.parse.quote(f"'{name}'!{values_range}")
    return _api("GET", f"{SHEETS_API}/{SHEET_ID}/values/{rng}?majorDimension=ROWS")


def sheets_update(values_range, values):
    name, _ = find_main_tab()
    if not name:
        raise RuntimeError("ไม่พบ tab 台帳 หลัก")
    rng = urllib.parse.quote(f"'{name}'!{values_range}")
    return _api("PUT", f"{SHEETS_API}/{SHEET_ID}/values/{rng}?valueInputOption=USER_ENTERED",
                body={"values": values})


def batch_update(requests, gid=None):
    return _api("POST", f"{SHEETS_API}/{SHEET_ID}:batchUpdate", body={"requests": requests})


def month_sections(rows, base_row):
    sections = {}
    for i, row in enumerate(rows):
        a = row[0] if row else None
        if a and isinstance(a, str):
            m = re.search(r"(\d+)月", a)
            if m:
                sections[int(m.group(1))] = base_row + i + 3
    return sections


def col_for_day(day):
    n = 1 + day
    if n <= 26:
        return chr(ord("A") + n - 1)
    return "A" + chr(ord("A") + n - 27)


def read_taicho():
    res = sheets_get("A1:AG80")
    rows = res.get("values", [])
    sections = month_sections(rows, 1)
    taicho = {}
    for month, entry_row in sections.items():
        idx = entry_row - 1
        entry = rows[idx] if idx < len(rows) else []
        cells = {}
        for day in range(1, 32):
            c = col_for_day(day)
            ci = ord(c[0]) - ord("A") if len(c) == 1 else (ord(c[0]) - ord("A")) * 26 + ord(c[1]) - ord("A") + 26
            v = entry[ci] if ci < len(entry) else None
            if v not in (None, ""):
                cells[day] = v
        ndays = DAYS_IN_MONTH.get(month, 30)
        hidx = idx - 3
        if hidx >= 0:
            hdr = rows[hidx] if hidx < len(rows) else []
            nums = []
            for v in hdr[1:]:
                try:
                    n = int(v)
                    if 1 <= n <= 31:
                        nums.append(n)
                except (TypeError, ValueError):
                    pass
            if nums:
                ndays = max(nums)
        taicho[month] = {"row": entry_row, "cells": cells, "ndays": ndays}
    return taicho


# ---------------- ใบขอรถ ----------------

def serial_to_date(n):
    if isinstance(n, _dt.datetime):
        return n.date()
    if isinstance(n, (int, float)):
        return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=n)).date()
    return None


def fmt_time(c):
    if isinstance(c, _dt.time):
        return c.strftime("%H:%M")
    if isinstance(c, str):
        if "→" in c:
            c = c.split("→")[-1].strip()
        c = re.sub(r"\s+", "", c)
        if re.match(r"^\d{1,2}[:：]\d{2}$", c):
            hh, mm = re.split(r"[:：]", c)
            return f"{int(hh):02d}:{mm}"
    return None


def hours_from(keiyu):
    if keiyu:
        m = re.search(r"\((\d+)H\)", str(keiyu))
        if m:
            return f"{m.group(1)}H"
    return None


def parse_form_bytes(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    out = {}
    for tab, month in TAB_MONTHS.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(10, ws.max_row + 1):
            n = ws.cell(row=r, column=14).value
            d = serial_to_date(n)
            if d is None or d.month != month or d.year not in (2026, 2027):
                continue
            j = ws.cell(row=r, column=10).value
            if j and "キャンセル" in str(j):
                continue
            c = fmt_time(ws.cell(row=r, column=3).value)
            if c is None:
                continue
            e = ws.cell(row=r, column=5).value
            f = ws.cell(row=r, column=6).value
            h = hours_from(e) or hours_from(f)
            entry = f"🔷{c} {h}" if h else c
            out.setdefault(d, []).append(entry)
    def _tk(e):
        m = re.search(r"(\d{2}):(\d{2})", e)
        return m.groups() if m else ("99", "99")
    return {d: sorted(v, key=_tk) for d, v in out.items()}


def strip_flags(s):
    return re.sub(r"[🟡🟢]", "", s or "").rstrip("\r\n")


def build_plan(form_rows):
    taicho = read_taicho()
    plan = {}
    for d, entries in sorted(form_rows.items()):
        month = d.month
        if month not in taicho:
            continue
        target = "\n".join(entries)
        current = taicho[month]["cells"].get(d.day)
        cur_str = str(current) if current is not None else None
        if strip_flags(cur_str) != target:
            reason = "ลบ (งานยกเลิก/ไม่มีแล้ว)" if target == "" else (
                "ลงใหม่" if cur_str is None else "แก้ไข")
            plan.setdefault(month, {})[d.day] = (target, reason, cur_str)
    return plan


def apply_plan(plan, dry_run=True):
    taicho = read_taicho()
    for month, days in sorted(plan.items()):
        entry_row = taicho[month]["row"]
        for day in sorted(days):
            target, reason, cur = days[day]
            col = col_for_day(day)
            rng = f"{col}{entry_row}"
            write_val = (target + "🟢") if target else target
            if dry_run:
                print(f"  [{reason}] {month}月{day:02d} {rng}: {cur!r} -> {write_val!r}", flush=True)
            else:
                sheets_update(rng, [[write_val]])
                print(f"  [เขียน] {month}月{day:02d} {rng}: {cur!r} -> {write_val!r}", flush=True)


# ---------------- flags / rotate ----------------

def clear_stale_flags(state):
    today = date.today().isoformat()
    if state.get("last_flag_date") == today:
        return 0
    if "last_flag_date" not in state:
        state["last_flag_date"] = today
        print(f"flag bootstrap: last_flag_date={today}", flush=True)
        return 0
    t = read_taicho()
    n = 0
    for month, info in t.items():
        row = info["row"]
        for day, v in info["cells"].items():
            if "🟢" in v or "🟡" in v:
                clean = strip_flags(v)
                if clean != v:
                    sheets_update(f"{col_for_day(day)}{row}", [[clean]])
                    n += 1
    state["last_flag_date"] = today
    print(f"cleared {n} stale flag(s)", flush=True)
    return n


def rotate_table_if_needed():
    if date.today().day != 1:
        return False
    rotated = False
    while True:
        name, gid = find_main_tab()
        if not name:
            break
        t = read_taicho()
        if not t:
            break
        months = sorted(t.keys())
        first, last = months[0], months[-1]
        if first == date.today().month:
            break
        rotated = True
        first_entry = t[first]["row"]
        first_header = first_entry - 3
        last_header = t[last]["row"] - 3
        new_month = (last % 12) + 1
        new_year = 2027 if new_month == 1 else 2026
        year_label = 2027 if new_month < last else 2026
        m0 = (first % 12) + 1
        m3 = new_month
        new_title = f"千栄1568 {m0}月-{m3}月({year_label})"
        second_header = (t[months[1]]["row"] - 3) if len(months) > 1 else first_header + MONTH_BLOCK
        first_block = second_header - first_header
        prev_header = (t[months[-2]]["row"] - 3) if len(months) > 1 else last_header - MONTH_BLOCK
        last_block = last_header - prev_header
        tmp_title = "千栄1568 tmp"
        resp = _api("POST", f"{SHEETS_API}/{SHEET_ID}:batchUpdate",
                    body={"requests": [{"duplicateSheet": {
                        "sourceSheetId": gid, "insertSheetIndex": 0,
                        "newSheetName": tmp_title}}]})
        new_gid = None
        for r in resp.get("replies", []):
            p = r.get("duplicateSheet", {}).get("properties", {})
            new_gid = p.get("sheetId")
        if new_gid is None:
            raise RuntimeError("duplicateSheet failed")
        batch_update([{"deleteDimension": {"range": {"sheetId": new_gid, "dimension": "ROWS",
                        "startIndex": first_header - 1, "endIndex": first_header - 1 + first_block}}}])
        tok = sheets_token()
        a_rows = _api("GET", f"{SHEETS_API}/{SHEET_ID}/values/A1:A80?majorDimension=ROWS").get("values", [])
        last_hdr_1based = None
        for i, row in enumerate(a_rows, start=1):
            a = row[0] if row else ""
            if re.search(r"\d+月", str(a)):
                last_hdr_1based = i
        if last_hdr_1based is None:
            raise RuntimeError("หา header ล่าสุดไม่เจอ")
        dest = last_hdr_1based - 1 + last_block
        src = dest - last_block
        batch_update([{"copyPaste": {"source": {"sheetId": new_gid, "startRowIndex": src,
                        "endRowIndex": dest, "startColumnIndex": 0, "endColumnIndex": 33},
                        "destination": {"sheetId": new_gid, "startRowIndex": dest, "endColumnIndex": 33},
                        "pasteType": "PASTE_FORMAT"}}])
        h = dest + 1
        days = list(range(1, 32))
        rng = f"A{h}:AF{h + 9}"
        quote_rng = urllib.parse.quote(rng)
        _api("PUT", f"{SHEETS_API}/{SHEET_ID}/values/{quote_rng}?valueInputOption=USER_ENTERED",
             body={"values": [[f"{new_year}年 {new_month}月"] + days, [None] * 33, [None] * 33,
                              ["千栄1568"] + [None] * 32, [None] * 33, [None] + days, [None] * 33,
                              ["変更は🟡"] + [None] * 32, ["新規は🟢"] + [None] * 32,
                              ["経由地🔷"] + [None] * 32]})
        batch_update([{"updateSheetProperties": {"properties": {"sheetId": new_gid, "title": new_title},
                                                 "fields": "title"}}])
        reqs = [
            {"updateDimensionProperties": {"range": {"sheetId": new_gid, "dimension": "ROWS",
                        "startIndex": 0, "endIndex": 956},
                        "properties": {"hiddenByUser": False}, "fields": "hiddenByUser"}},
            {"updateDimensionProperties": {"range": {"sheetId": new_gid, "dimension": "ROWS",
                        "startIndex": h + 2, "endIndex": h + 3},
                        "properties": {"pixelSize": 300}, "fields": "pixelSize"}},
            {"repeatCell": {"range": {"sheetId": new_gid, "startRowIndex": h + 10, "endRowIndex": h + 11,
                        "startColumnIndex": 0, "endColumnIndex": 33},
                        "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.4, "green": 0.4, "blue": 0.4}}},
                        "fields": "userEnteredFormat.backgroundColor"}},
            {"repeatCell": {"range": {"sheetId": new_gid, "startRowIndex": h + 13, "endRowIndex": h + 14,
                        "startColumnIndex": 0, "endColumnIndex": 33},
                        "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.4, "green": 0.4, "blue": 0.4}}},
                        "fields": "userEnteredFormat.backgroundColor"}},
            {"updateDimensionProperties": {"range": {"sheetId": new_gid, "dimension": "ROWS",
                        "startIndex": h + 15, "endIndex": 956},
                        "properties": {"hiddenByUser": True}, "fields": "hiddenByUser"}},
        ]
        t_new = read_taicho()
        for m, info in t_new.items():
            hdr = info["row"] - 3
            r, g, b = HEADER_COLOR.get(m, (0.9, 0.9, 0.9))
            reqs.append({"repeatCell": {"range": {"sheetId": new_gid, "startRowIndex": hdr - 1,
                        "endRowIndex": hdr, "startColumnIndex": 0, "endColumnIndex": 33},
                        "cell": {"userEnteredFormat": {"backgroundColor": {"red": r, "green": g, "blue": b}}},
                        "fields": "userEnteredFormat.backgroundColor"}})
        batch_update(reqs)
        print(f"rotate: {new_title}", flush=True)
    return rotated


# ---------------- orchestrator ----------------

def run_flow():
    token = sheets_token()
    state = load_state(token)
    try:
        rotate_table_if_needed()
        clear_stale_flags(state)
    except Exception as e:
        print(f"rotate/clear failed: {e}", flush=True)
    forms = list_forms(token)
    if state.get("processed") is None:
        state["processed"] = []
    if not state["processed"] and forms:
        newest = forms[-1][2]
        for _, _, nm in forms:
            if nm != newest:
                state["processed"].append(nm)
        print(f"bootstrap: marked {len(state['processed'])} old forms", flush=True)
    new_forms = [(mmdd, fid, nm) for mmdd, fid, nm in forms if nm not in state["processed"]]
    if not new_forms:
        print("No new form. done.", flush=True)
        save_state(state, token)
        return "no new form"
    for mmdd, fid, nm in new_forms:
        try:
            data = _get_media(f"{DRIVE_API}/files/{fid}?alt=media", token)
            plan = build_plan(parse_form_bytes(data))
            if not plan:
                msg = f"✅ 台帳を確認しました（{mmdd}）: 変更なし"
                print("no change:", nm, flush=True)
                send_line(f"{msg}\n{SHORT_TAICHO_URL}")
                send_line("宜しくお願い致します。")
                state["processed"].append(nm)
                save_state(state, token)
                continue
            apply_plan(plan, dry_run=False)
            JA_REASON = {"ลงใหม่": "新規🟢", "แก้ไข": "変更🟡", "ลบ (งานยกเลิก/ไม่มีแล้ว)": "削除"}
            lines = []
            total = 0
            for month in sorted(plan):
                for day in sorted(plan[month]):
                    target, reason, cur = plan[month][day]
                    disp = (target + "🟢") if target else target
                    lines.append(f"  - {month}月{day}日 [{JA_REASON.get(reason, reason)}]: {disp!r}")
                    total += 1
            msg = (f"📋 台帳を自動更新しました（{mmdd}、{total}箇所）:\n" + "\n".join(lines))
            send_line(msg)
            send_line(f"📄 台帳を開く（{mmdd}）\n{SHORT_TAICHO_URL}")
            try:
                archive_pdf(mmdd, token)  # phase 2b: PDF reportlab -> Drive โฟลเดอร์เดือน
            except Exception as e:
                print(f"PDF archive failed ({mmdd}): {e}", flush=True)
            send_line("宜しくお願い致します。")
            state["processed"].append(nm)
            save_state(state, token)
        except Exception as e:
            print(f"failed {nm}: {e}", flush=True)
            try:
                send_line(f"⚠️ 台帳更新に失敗（{mmdd}）: {e}")
            except Exception:
                pass
    save_state(state, token)
    return f"processed {len(new_forms)} form(s)"


@https_fn.on_request(region=SupportedRegion.ASIA_SOUTHEAST1)
def taicho_monitor(req: https_fn.Request) -> https_fn.Response:
    """HTTP entry. body/query {pdf: 1, mmdd?: MMDD} = สร้าง/อัปเดต PDF archive เท่านั้น
    (ไม่แตะ sheets/LINE/state — ใช้ตรวจสอบหรือสร้าง PDF ใหม่ด้วยมือ)"""
    try:
        body = {}
        if req.method == "POST":
            raw = req.get_data(as_text=True)
            if raw and raw.lstrip()[:1] in ("{", "["):
                body = json.loads(raw)
        body.update({k: v for k, v in (req.args.items() if req.args else [])})
        if body.get("pdf") == "1":
            token = sheets_token()
            mmdd = str(body.get("mmdd") or _dt.date.today().strftime("%m%d"))
            archive_pdf(mmdd, token)
            return https_fn.Response(f"PDF archived {mmdd}", status=200)
        msg = run_flow()
        return https_fn.Response(f"OK {msg}", status=200)
    except Exception as e:
        print(f"ERROR: {e}", flush=True)
        return https_fn.Response(f"ERROR: {e}", status=500)
